"""Tests fuer den XLSX-Export gegen das CATS-Mass-Upload-Muster."""
import io
import re
import sys
import zipfile
from datetime import date
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

import openpyxl  # noqa: E402

from xlsx_export import COLUMNS, ExportRow, build_bytes, suggest_filename  # noqa: E402

ROWS = [
    ExportRow(date(2026, 8, 17), "00123456", "DEO2222-NP/PJ00-O51.0000", 5.55),
    ExportRow(date(2026, 8, 17), "00123456", "DEO1111-NP/PJ00-O51.0000", 2.0),
]


def sheet():
    return openpyxl.load_workbook(io.BytesIO(build_bytes(ROWS))).active


def test_kopf_steht_dreimal_daten_ab_zeile_vier():
    ws = sheet()
    for row in (1, 2, 3):
        assert [c.value for c in ws[row]] == COLUMNS
    assert ws["A4"].value is not None


def test_workdate_ist_echtes_datum_mit_builtin_format():
    ws = sheet()
    assert ws["A4"].value.date() == date(2026, 8, 17)
    assert ws["A4"].number_format == "mm-dd-yy"      # numFmtId=14


def test_datum_wird_als_serial_gespeichert():
    # Das Muster enthaelt 46251 fuer den 17.08.2026.
    z = zipfile.ZipFile(io.BytesIO(build_bytes(ROWS)))
    xml = z.read("xl/worksheets/sheet1.xml").decode()
    assert "46251" in xml
    styles = z.read("xl/styles.xml").decode()
    assert 'numFmtId="14"' in re.search(r"<cellXfs.*?</cellXfs>", styles, re.S).group(0)


def test_employee_ohne_fuehrende_nullen_als_zahl():
    ws = sheet()
    assert ws["B4"].value == 123456
    assert isinstance(ws["B4"].value, int)


def test_zeilen_nach_datum_und_wbs_sortiert():
    ws = sheet()
    assert ws["C4"].value == "DEO1111-NP/PJ00-O51.0000"
    assert ws["C5"].value == "DEO2222-NP/PJ00-O51.0000"


def test_optionale_spalten_bleiben_leer():
    ws = sheet()
    for col in ("D", "E", "F", "G", "I", "J"):
        assert ws[f"{col}4"].value is None


def test_stunden_als_zahl():
    ws = sheet()
    assert ws["H4"].value == 2.0
    assert ws["H5"].value == 5.55


def test_nicht_numerische_personalnummer_bleibt_text():
    rows = [ExportRow(date(2026, 8, 17), "AB1234", "DEO1111-NP/PJ00-O51.0000", 1.0)]
    ws = openpyxl.load_workbook(io.BytesIO(build_bytes(rows))).active
    assert ws["B4"].value == "AB1234"


def test_leerer_export_hat_nur_die_kopfzeilen():
    ws = openpyxl.load_workbook(io.BytesIO(build_bytes([]))).active
    assert ws.max_row == 3


def test_dateiname():
    assert suggest_filename(date(2026, 8, 1), date(2026, 8, 17)) == "CATS_20260801_20260817.xlsx"
