"""Erzeugt die CATS-Mass-Upload-XLSX.

Das Zielformat folgt exakt dem Muster (SPEC.md §6):

  * ein Sheet, Spalten A-J
  * die Kopfzeile steht dreimal (Zeile 1, 2, 3), Daten ab Zeile 4
  * WORKDATE  als echtes Excel-Datum, Zahlenformat numFmtId=14
  * EMPLOYEE  als Zahl ohne fuehrende Nullen
  * CATSHOURS als Zahl mit zwei Nachkommastellen
  * ABS_ATT_TYPE, REC_ORDER, ACTIVITY, WAGETYPE, ZZICTPC, SHORTTEXT bleiben leer
"""
from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

COLUMNS = [
    "WORKDATE",
    "EMPLOYEE",
    "WBS_ELEMENT",
    "ABS_ATT_TYPE",
    "REC_ORDER",
    "ACTIVITY",
    "WAGETYPE",
    "CATSHOURS",
    "ZZICTPC",
    "SHORTTEXT",
]

HEADER_ROWS = 3          # das Muster wiederholt den Kopf dreimal
DATE_FORMAT = "mm-dd-yy"  # exakt der Builtin numFmtId=14 (wie im Muster)
HOURS_FORMAT = "0.00"

# Spaltenbreiten aus dem Muster
COLUMN_WIDTHS = {
    "A": 11.18, "B": 11.45, "C": 24.82, "D": 12.54, "E": 12.54,
    "F": 11.27, "G": 11.27, "I": 15.73,
}


@dataclass(frozen=True)
class ExportRow:
    work_date: date
    employee: str          # Personalnummer, fuehrende Nullen erlaubt
    wbs_element: str
    hours: float


def _employee_number(raw: str) -> int | str:
    """'00123456' -> 123456. Nicht-numerische Werte bleiben unveraendert."""
    stripped = str(raw).strip()
    return int(stripped) if stripped.isdigit() else stripped


def build_workbook(rows: list[ExportRow]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    for _ in range(HEADER_ROWS):
        ws.append(COLUMNS)

    for row in sorted(rows, key=lambda r: (r.work_date, r.wbs_element)):
        ws.append([
            row.work_date,
            _employee_number(row.employee),
            row.wbs_element,
            None, None, None, None,
            round(float(row.hours), 2),
            None, None,
        ])

    for cell in ws[f"A{HEADER_ROWS + 1}":f"A{ws.max_row}"] if ws.max_row > HEADER_ROWS else []:
        cell[0].number_format = DATE_FORMAT
    for r in range(HEADER_ROWS + 1, ws.max_row + 1):
        ws.cell(row=r, column=8).number_format = HOURS_FORMAT

    for col, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col].width = width

    return wb


def build_bytes(rows: list[ExportRow]) -> bytes:
    buf = BytesIO()
    build_workbook(rows).save(buf)
    return buf.getvalue()


def suggest_filename(date_from: date, date_to: date) -> str:
    return f"CATS_{date_from:%Y%m%d}_{date_to:%Y%m%d}.xlsx"
